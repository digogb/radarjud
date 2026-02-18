"""
Serviço de importação de pessoas monitoradas a partir de planilha Excel.

Lê o arquivo pessoas.xlsx e cadastra as partes adversas como PessoaMonitorada.
A data de expiração é calculada como data_prazo + 5 anos.
"""

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from dateutil.relativedelta import relativedelta

from storage.repository import DiarioRepository

logger = logging.getLogger(__name__)

ANOS_MONITORAMENTO = 5

# Headers esperados na planilha (busca por nome para ser robusto a reordenações)
HEADER_DATA_PRAZO = "Data Prazo"
HEADER_PARTE_ADVERSA = "Parte Adversa"
HEADER_CPF_CNPJ = "CPF/CNPJ Adverso"
HEADER_NUMERO_PROCESSO = "Número do Processo"
HEADER_COMARCA = "Comarca"
HEADER_UF = "UF"


def extrair_nome(parte_adversa: str) -> Optional[str]:
    """Extrai o nome a partir do campo 'Parte Adversa'.

    Exemplos:
        'AUTOR - LUCIO MACHADO DA SILVA'  → 'LUCIO MACHADO DA SILVA'
        ' - JANE MARY ABUHASSAN GONÇALVES' → 'JANE MARY ABUHASSAN GONÇALVES'
        'MARIA SILVA'                      → 'MARIA SILVA'
    """
    if not parte_adversa:
        return None
    # Split no primeiro ' - ' e pega o que vem depois
    partes = str(parte_adversa).split(" - ", 1)
    nome = partes[1].strip() if len(partes) == 2 else partes[0].strip()
    return nome.upper() if nome else None


def normalizar_cpf(raw) -> Optional[str]:
    """Remove formatação do CPF/CNPJ e retorna apenas os dígitos.

    CPF válido: 11 dígitos. CNPJ: 14 dígitos. Outros: retorna None.
    """
    if not raw:
        return None
    digitos = re.sub(r"\D", "", str(raw))
    if len(digitos) in (11, 14):
        return digitos
    return None


def parse_data_prazo(raw) -> Optional[date]:
    """Converte string de data 'dd/mm/yyyy HH:MM:SS' ou 'dd/mm/yyyy' para date."""
    if not raw:
        return None
    s = str(raw).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _mapear_colunas(ws) -> dict:
    """Retorna mapeamento {nome_header: índice_coluna_0based} a partir da linha 1."""
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column - 1
    return headers


def importar_planilha(
    filepath: str,
    repo: DiarioRepository,
    dry_run: bool = False,
) -> dict:
    """Lê a planilha e cadastra partes adversas como pessoas monitoradas.

    Args:
        filepath: Caminho para o arquivo .xlsx
        repo: Repositório do banco de dados
        dry_run: Se True, apenas valida sem gravar no banco

    Returns:
        Dict com estatísticas: total, importados, atualizados, pulados, erros
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    col = _mapear_colunas(ws)

    # Validar colunas obrigatórias
    for header in (HEADER_PARTE_ADVERSA,):
        if header not in col:
            raise ValueError(f"Coluna obrigatória não encontrada na planilha: '{header}'")

    stats = {"total": 0, "importados": 0, "pulados": 0, "erros": 0}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        stats["total"] += 1
        try:
            parte_adversa = row[col[HEADER_PARTE_ADVERSA]] if HEADER_PARTE_ADVERSA in col else None
            nome = extrair_nome(parte_adversa)

            if not nome:
                logger.debug(f"Linha {row_idx}: Parte Adversa vazia ou inválida — pulando")
                stats["pulados"] += 1
                continue

            cpf_raw = row[col[HEADER_CPF_CNPJ]] if HEADER_CPF_CNPJ in col else None
            cpf = normalizar_cpf(cpf_raw)

            data_prazo_raw = row[col[HEADER_DATA_PRAZO]] if HEADER_DATA_PRAZO in col else None
            dt_prazo = parse_data_prazo(data_prazo_raw)
            dt_expiracao = (dt_prazo + relativedelta(years=ANOS_MONITORAMENTO)) if dt_prazo else None

            numero_processo_raw = row[col[HEADER_NUMERO_PROCESSO]] if HEADER_NUMERO_PROCESSO in col else None
            numero_processo = str(numero_processo_raw).strip() if numero_processo_raw else None

            comarca_raw = row[col[HEADER_COMARCA]] if HEADER_COMARCA in col else None
            comarca = str(comarca_raw).strip() if comarca_raw else None

            uf_raw = row[col[HEADER_UF]] if HEADER_UF in col else None
            uf = str(uf_raw).strip().upper() if uf_raw else None

            if dry_run:
                logger.info(
                    f"[DRY RUN] Linha {row_idx}: {nome} | CPF: {cpf} | "
                    f"Processo: {numero_processo} | Prazo: {dt_prazo} | Exp: {dt_expiracao}"
                )
                stats["importados"] += 1
                continue

            repo.adicionar_pessoa(
                nome=nome,
                cpf=cpf,
                numero_processo=numero_processo,
                comarca=comarca,
                uf=uf,
                data_prazo=dt_prazo,
                data_expiracao=dt_expiracao,
                origem_importacao="PLANILHA",
            )
            logger.info(f"Linha {row_idx}: importado — {nome}")
            stats["importados"] += 1

        except Exception as e:
            logger.error(f"Linha {row_idx}: erro ao importar — {e}", exc_info=True)
            stats["erros"] += 1

    wb.close()
    return stats
